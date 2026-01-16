"""Tabular file extraction (CSV, XLSX)."""

from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


def extract_text_from_csv(file_path: str) -> List[Dict]:
    """Extract text from CSV file."""
    try:
        import csv
        
        pages = []
        page_num = 1
        rows_per_page = 100
        current_page_rows = []
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            # Try to detect delimiter
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            reader = csv.reader(f, delimiter=delimiter)
            
            # Read header
            try:
                header = next(reader)
                current_page_rows.append(" | ".join(header))
            except StopIteration:
                pass
            
            # Read rows
            for row in reader:
                current_page_rows.append(" | ".join(row))
                
                if len(current_page_rows) >= rows_per_page:
                    pages.append({
                        "page_number": page_num,
                        "text": "\n".join(current_page_rows)
                    })
                    current_page_rows = []
                    page_num += 1
            
            # Add remaining rows
            if current_page_rows:
                pages.append({
                    "page_number": page_num,
                    "text": "\n".join(current_page_rows)
                })
        
        return pages if pages else [{"page_number": 1, "text": ""}]
        
    except Exception as e:
        logger.error(f"Error extracting text from CSV {file_path}: {str(e)}")
        raise Exception(f"Failed to extract text from CSV: {str(e)}")


def extract_text_from_xlsx(file_path: str) -> List[Dict]:
    """Extract text from XLSX file."""
    try:
        try:
            import openpyxl
            use_openpyxl = True
        except ImportError:
            try:
                import pandas as pd
                use_openpyxl = False
            except ImportError:
                raise ImportError("Neither openpyxl nor pandas is installed. Install one with: pip install openpyxl or pip install pandas")
        
        pages = []
        
        if use_openpyxl:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            page_num = 1
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_texts = []
                
                # Add sheet name as header
                sheet_texts.append(f"Sheet: {sheet_name}")
                sheet_texts.append("")
                
                # Extract data from cells
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        sheet_texts.append(row_text)
                
                if sheet_texts:
                    pages.append({
                        "page_number": page_num,
                        "text": "\n".join(sheet_texts)
                    })
                    page_num += 1
            
            wb.close()
        else:
            # Use pandas
            df_dict = pd.read_excel(file_path, sheet_name=None)
            page_num = 1
            
            for sheet_name, df in df_dict.items():
                sheet_texts = [f"Sheet: {sheet_name}", ""]
                sheet_texts.append(df.to_string())
                
                pages.append({
                    "page_number": page_num,
                    "text": "\n".join(sheet_texts)
                })
                page_num += 1
        
        return pages if pages else [{"page_number": 1, "text": ""}]
        
    except Exception as e:
        logger.error(f"Error extracting text from XLSX {file_path}: {str(e)}")
        raise Exception(f"Failed to extract text from XLSX: {str(e)}")

